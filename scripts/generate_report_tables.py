"""
Generate Excel Report Tables - Complete Statistical Product
============================================================
Health Financing Gap in Africa - All Indicators from Specification

Includes:
- Gini coefficients for inequality measurement
- Count indicators (number of countries meeting thresholds)
- Cross-tabulation analyses (financing dimensions × outcomes)
- On-course calculations for mortality targets
- All disaggregations by income status, sub-region, and year
"""

import pandas as pd
import numpy as np
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Paths
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(BASE_DIR, 'processed_data')
OUTPUT_DIR = os.path.join(BASE_DIR, 'reports')

os.makedirs(OUTPUT_DIR, exist_ok=True)

# Styling
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SECTION_FILL = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
SECTION_FONT = Font(bold=True)
THEME_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
THEME_FONT = Font(color="FFFFFF", bold=True, size=12)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Thresholds
THRESHOLDS = {
    'Low': 112,
    'Lower-middle': 146,
    'Upper-middle': 477
}
ABUJA_TARGET = 15
OOP_BENCHMARK = 20
NMR_TARGET = 12
MMR_TARGET = 70
UHC_LOW_THRESHOLD = 50


def load_data():
    """Load and prepare the master dataset."""
    df = pd.read_csv(os.path.join(DATA_DIR, 'master_dataset.csv'))
    # Filter for 2000-2023 (excluding 2024)
    df = df[(df['year'] >= 2000) & (df['year'] <= 2023)]
    df = df[df['location'].notna()]

    # Recode 'High' income to 'Upper-middle' (only 1 country: Seychelles)
    df['income'] = df['income'].replace('High', 'Upper-middle')

    # Compute additional indicators needed for the specification
    df = compute_additional_indicators(df)

    return df


def compute_additional_indicators(df):
    """Compute additional indicators required by the specification."""

    # 3.1.1 Countries BELOW threshold (not meeting target - bad)
    # Spec: "countries with domestic general government health expenditure per capita BELOW the threshold"
    df['below_threshold'] = df['Gov exp Health per capita More than Threshold'].apply(
        lambda x: 0 if x == 1 else 1 if pd.notna(x) else np.nan
    )

    # 3.2.1 Countries BELOW Abuja target (< 15% - not meeting target)
    # Spec: "countries with share... BELOW the Abuja Declaration target (< 15%)"
    df['below_abuja'] = df['Gov exp Health on budget > 15'].apply(
        lambda x: 0 if x == 1 else 1 if pd.notna(x) else np.nan
    )

    # 3.3.1 Countries with OOP BELOW the 20% benchmark (good financial protection)
    # Spec: "countries with share of OOP... below the benchmark (> 20%)"
    # OOP < 20% means good financial protection (below the problematic threshold)
    df['below_oop_benchmark'] = df['Out-of-pocket on health exp < 20'].copy()

    # Also keep indicator for countries with OOP > 20% (poor protection) for other analyses
    df['oop_above_20'] = df['Out-of-pocket on health exp < 20'].apply(
        lambda x: 0 if x == 1 else 1 if pd.notna(x) else np.nan
    )

    # 3.4.1 Government share dominant (> 50%)
    df['gov_dominant'] = (df['Govern on health exp'] > 50).astype(float)
    df.loc[df['Govern on health exp'].isna(), 'gov_dominant'] = np.nan

    # 3.4a Countries with Government share of health expenditure < 5% of GDP
    df['gov_health_below_5pct_gdp'] = (df['Gov exp Health on GDP'] < 5).astype(float)
    df.loc[df['Gov exp Health on GDP'].isna(), 'gov_health_below_5pct_gdp'] = np.nan

    # 3.5.2 UHC below 50% OR below regional average
    # Spec: "UHC index of less than 50 per cent OR lower than the average for the region"
    df['uhc_below_50'] = (df['Universal health coverage'] < 50).astype(float)
    df.loc[df['Universal health coverage'].isna(), 'uhc_below_50'] = np.nan

    # Calculate UHC below regional average per year
    for year in df['year'].unique():
        year_mask = df['year'] == year
        year_avg = df.loc[year_mask, 'Universal health coverage'].mean()
        if pd.notna(year_avg):
            df.loc[year_mask, 'uhc_below_avg'] = (
                df.loc[year_mask, 'Universal health coverage'] < year_avg
            ).astype(float)
            df.loc[year_mask & df['Universal health coverage'].isna(), 'uhc_below_avg'] = np.nan

    # Combined: below 50% OR below average
    df['uhc_below_50_or_avg'] = ((df['uhc_below_50'] == 1) | (df['uhc_below_avg'] == 1)).astype(float)
    df.loc[df['Universal health coverage'].isna(), 'uhc_below_50_or_avg'] = np.nan

    # 3.6.3 On course for NMR target: "reduce NMR to at least 12 per 1,000"
    # "at least 12" means reaching 12 or below, so <= 12
    df['nmr_on_course'] = (df['Neonatal mortality rate'] <= NMR_TARGET).astype(float)
    df.loc[df['Neonatal mortality rate'].isna(), 'nmr_on_course'] = np.nan

    # 3.6.4 On course for MMR target: "reduce MMR to less than 70 per 100,000"
    # "less than 70" means < 70 (strict inequality)
    df['mmr_on_course'] = (df['Maternal mortality ratio'] < MMR_TARGET).astype(float)
    df.loc[df['Maternal mortality ratio'].isna(), 'mmr_on_course'] = np.nan

    # Identify dominant financing source for each country-year
    financing_cols = ['Govern on health exp', 'Voluntary Prepayments on health exp',
                      'Out-of-pocket on health exp', 'Other Private on health exp',
                      'External on health exp']

    def get_dominant_source(row):
        values = {col: row[col] for col in financing_cols if pd.notna(row.get(col))}
        if not values:
            return np.nan
        return max(values, key=values.get)

    df['dominant_financing_source'] = df.apply(get_dominant_source, axis=1)

    # UHC percentiles (50th and 75th) - calculated per year
    # Spec: "UHC score of more than the 50th and 75th percentiles for the continent"
    for year in df['year'].unique():
        year_data = df[df['year'] == year]['Universal health coverage'].dropna()
        if len(year_data) > 0:
            p50 = year_data.quantile(0.50)
            p75 = year_data.quantile(0.75)
            mask = df['year'] == year
            # "more than" means > (strict inequality)
            df.loc[mask, 'uhc_above_p50'] = (df.loc[mask, 'Universal health coverage'] > p50).astype(float)
            df.loc[mask, 'uhc_above_p75'] = (df.loc[mask, 'Universal health coverage'] > p75).astype(float)
            df.loc[mask & df['Universal health coverage'].isna(), 'uhc_above_p50'] = np.nan
            df.loc[mask & df['Universal health coverage'].isna(), 'uhc_above_p75'] = np.nan

    # === NEW: Threshold categories for cross-tabulations ===
    # Get income-specific threshold for each country
    def get_threshold(income):
        if income == 'Low':
            return 112
        elif income == 'Lower-middle':
            return 146
        elif income == 'Upper-middle' or income == 'High':
            return 477
        else:
            return np.nan

    df['income_threshold'] = df['income'].apply(get_threshold)
    df['thres50'] = df['income_threshold'] * 0.5
    df['thres75'] = df['income_threshold'] * 0.749

    # Categorize gov expenditure relative to threshold
    def categorize_exp(row):
        gov_exp = row['Gov exp Health per capita']
        thres50 = row['thres50']
        thres75 = row['thres75']
        threshold = row['income_threshold']

        if pd.isna(gov_exp) or pd.isna(threshold):
            return np.nan
        elif gov_exp < thres50:
            return 'Below 50% of expenditure threshold'
        elif gov_exp >= thres50 and gov_exp < thres75:
            return '50-74.9% of expenditure threshold'
        elif gov_exp >= thres75 and gov_exp < threshold:
            return '75-99.9% of expenditure threshold'
        else:
            return 'Meet expenditure threshold target'

    df['gov_exp_pc_thres'] = df.apply(categorize_exp, axis=1)

    # UHC above 75%
    df['uhc_above_50'] = df['Universal health coverage'].apply(
        lambda x: '>50%' if pd.notna(x) and x > 50 else ('<=50%' if pd.notna(x) else np.nan)
    )

    # NMR categories (>12 or <=12)
    df['nmr_category'] = df['Neonatal mortality rate'].apply(
        lambda x: '>12' if pd.notna(x) and x > 12 else ('<=12' if pd.notna(x) else np.nan)
    )

    # MMR categories (>70 or <=70)
    df['mmr_category'] = df['Maternal mortality ratio'].apply(
        lambda x: '>70' if pd.notna(x) and x > 70 else ('<=70' if pd.notna(x) else np.nan)
    )

    return df


def calculate_gini(values):
    """Calculate Gini coefficient for a series of values."""
    values = np.array(values)
    values = values[~np.isnan(values)]

    if len(values) < 2:
        return np.nan

    values = np.sort(values)
    n = len(values)

    # Handle zero or negative values
    if values.sum() == 0:
        return np.nan

    # Gini formula
    index = np.arange(1, n + 1)
    gini = (2 * np.sum(index * values) - (n + 1) * np.sum(values)) / (n * np.sum(values))

    return gini


def calculate_range(values):
    """Calculate range (max - min) for a series."""
    values = np.array(values)
    values = values[~np.isnan(values)]

    if len(values) < 2:
        return np.nan

    return np.max(values) - np.min(values)


def aggregate_with_gini(df, indicator, group_by=None):
    """Calculate aggregates including Gini coefficient."""
    if indicator not in df.columns:
        return None

    def agg_func(x):
        valid = x.dropna()
        return pd.Series({
            'Average': valid.mean(),
            'Count': len(valid),
            'Min': valid.min(),
            'Max': valid.max(),
            'Range': calculate_range(valid),
            'Gini': calculate_gini(valid)
        })

    if group_by:
        result = df.groupby(['year', group_by])[indicator].apply(agg_func).unstack()
        result = result.reset_index()
        return result
    else:
        result = df.groupby('year')[indicator].apply(agg_func).unstack()
        result = result.reset_index()
        result.columns = ['Year', 'Average', 'Count', 'Min', 'Max', 'Range', 'Gini']
        return result


def count_by_condition(df, condition_col, group_by=None):
    """Count countries meeting a condition."""
    if condition_col not in df.columns:
        return None

    def count_func(x):
        valid = x.dropna()
        return pd.Series({
            'Total Countries': len(valid),
            'Meeting Condition': valid.sum(),
            'Not Meeting': len(valid) - valid.sum(),
            'Percentage Meeting': (valid.sum() / len(valid) * 100) if len(valid) > 0 else np.nan
        })

    if group_by:
        result = df.groupby(['year', group_by])[condition_col].apply(count_func).unstack()
        result = result.reset_index()
        return result
    else:
        result = df.groupby('year')[condition_col].apply(count_func).unstack()
        result = result.reset_index()
        result.columns = ['Year', 'Total Countries', 'Meeting Condition', 'Not Meeting', 'Percentage Meeting']
        return result


def cross_tabulation(df, condition1_col, condition2_col, group_by=None):
    """Cross-tabulate two conditions."""
    if condition1_col not in df.columns or condition2_col not in df.columns:
        return None

    def cross_func(sub_df):
        valid = sub_df[[condition1_col, condition2_col]].dropna()
        if len(valid) == 0:
            return pd.Series({
                'Total': 0,
                'Both Conditions': 0,
                'Only First': 0,
                'Only Second': 0,
                'Neither': 0,
                'Proportion Both': np.nan
            })

        both = ((valid[condition1_col] == 1) & (valid[condition2_col] == 1)).sum()
        only_first = ((valid[condition1_col] == 1) & (valid[condition2_col] == 0)).sum()
        only_second = ((valid[condition1_col] == 0) & (valid[condition2_col] == 1)).sum()
        neither = ((valid[condition1_col] == 0) & (valid[condition2_col] == 0)).sum()

        return pd.Series({
            'Total': len(valid),
            'Both Conditions': both,
            'Only First': only_first,
            'Only Second': only_second,
            'Neither': neither,
            'Proportion Both (%)': (both / len(valid) * 100) if len(valid) > 0 else np.nan
        })

    if group_by:
        result = df.groupby(['year', group_by]).apply(cross_func).reset_index()
        return result
    else:
        result = df.groupby('year').apply(cross_func).reset_index()
        return result


def add_dataframe_to_sheet(ws, df, start_row, title=None, is_count_table=False):
    """Add a DataFrame to a worksheet.

    Args:
        ws: Worksheet to add data to
        df: DataFrame to add
        start_row: Starting row number
        title: Optional title for the table
        is_count_table: If True, format all numeric values as integers (0 decimal places)
    """
    if df is None or len(df) == 0:
        return start_row + 1

    current_row = start_row

    if title:
        ws.cell(row=current_row, column=1, value=title)
        ws.cell(row=current_row, column=1).font = SECTION_FONT
        ws.cell(row=current_row, column=1).fill = SECTION_FILL
        for col in range(1, len(df.columns) + 1):
            ws.cell(row=current_row, column=col).fill = SECTION_FILL
            ws.cell(row=current_row, column=col).border = THIN_BORDER
        current_row += 1

    # Headers
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=str(col_name))
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER
    current_row += 1

    # Data
    # Columns that should have zero decimal places (counts)
    count_keywords = ['Count', 'Total Countries', 'Meeting Condition', 'Not Meeting',
                      'Both Conditions', 'Only First', 'Only Second', 'Neither']

    for row_idx, row_data in df.iterrows():
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col_idx)
            col_name = str(df.columns[col_idx - 1])

            if pd.isna(value):
                cell.value = None
            elif isinstance(value, (int, float)):
                # Check column type for formatting
                is_year_col = col_name.lower() == 'year'
                is_percentage_col = 'Percentage' in col_name or 'Proportion' in col_name
                is_count_col = any(kw in col_name for kw in count_keywords)

                if is_year_col:
                    # Year: integer without thousands separator
                    cell.value = int(round(value))
                    cell.number_format = '0'
                elif is_percentage_col:
                    # Percentages: 1 decimal place (check this BEFORE count check)
                    cell.value = round(value, 1)
                    cell.number_format = '0.0'
                elif is_count_table:
                    # Count table: all numeric values as integers
                    cell.value = int(round(value))
                    cell.number_format = '#,##0'
                elif is_count_col:
                    # Counts: integer with thousands separator
                    cell.value = int(round(value))
                    cell.number_format = '#,##0'
                else:
                    # All other numerics (averages, etc.): 1 decimal place
                    cell.value = round(value, 1)
                    cell.number_format = '0.0'
            else:
                cell.value = value
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center')
        current_row += 1

    return current_row + 1


def create_section_31(wb, df):
    """3.1 Adequacy of Public Health Financing Gap."""
    ws = wb.create_sheet(title="3.1 Public Health Financing")
    row = 1

    ws.cell(row=row, column=1, value="3.1 Adequacy of Public Health Financing Gap")
    ws.cell(row=row, column=1).font = Font(bold=True, size=14, color="1F4E79")
    row += 2

    # 3.1.1 Number below threshold
    ws.cell(row=row, column=1, value="3.1.1 Countries Below Internationally Prescribed Threshold")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="2E75B6")
    row += 1
    ws.cell(row=row, column=1, value="Thresholds: LICs=$112, LMICs=$146, UMICs=$477 per capita")
    ws.cell(row=row, column=1).font = Font(italic=True, size=10)
    row += 1

    agg = count_by_condition(df, 'below_threshold')
    row = add_dataframe_to_sheet(ws, agg, row, "Africa Level - Annual Counts")

    for group in ['income', 'Subregion']:
        agg = count_by_condition(df, 'below_threshold', group)
        row = add_dataframe_to_sheet(ws, agg, row, f"By {group} - Annual Counts")

    # 3.1.2 Average financing gap
    ws.cell(row=row, column=1, value="3.1.2 Average Public Health Financing Gap")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="2E75B6")
    row += 1

    agg = aggregate_with_gini(df, 'Gap for Gov exp Health per capita')
    row = add_dataframe_to_sheet(ws, agg, row, "Africa Level - Annual Statistics")

    for group in ['income', 'Subregion']:
        agg = aggregate_with_gini(df, 'Gap for Gov exp Health per capita', group)
        row = add_dataframe_to_sheet(ws, agg, row, f"By {group}")

    # 3.1.3 Gini coefficient
    ws.cell(row=row, column=1, value="3.1.3 Inequality in Government Health Expenditure Per Capita")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="2E75B6")
    row += 1

    agg = aggregate_with_gini(df, 'Gov exp Health per capita')
    row = add_dataframe_to_sheet(ws, agg, row, "Africa Level - Gini & Range")

    for group in ['income', 'Subregion']:
        agg = aggregate_with_gini(df, 'Gov exp Health per capita', group)
        row = add_dataframe_to_sheet(ws, agg, row, f"By {group}")

    adjust_column_widths(ws)


def create_section_32(wb, df):
    """3.2 Budgetary Priority Assigned to Health (Abuja Declaration)."""
    ws = wb.create_sheet(title="3.2 Abuja Declaration")
    row = 1

    ws.cell(row=row, column=1, value="3.2 Budgetary Priority Assigned to Health")
    ws.cell(row=row, column=1).font = Font(bold=True, size=14, color="1F4E79")
    row += 2

    # 3.2.1 Number below Abuja target
    ws.cell(row=row, column=1, value="3.2.1 Countries Below Abuja Declaration Target (< 15%)")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="2E75B6")
    row += 1

    agg = count_by_condition(df, 'below_abuja')
    row = add_dataframe_to_sheet(ws, agg, row, "Africa Level")

    for group in ['income', 'Subregion']:
        agg = count_by_condition(df, 'below_abuja', group)
        row = add_dataframe_to_sheet(ws, agg, row, f"By {group}")

    # 3.2.2 Average budget priority gap
    ws.cell(row=row, column=1, value="3.2.2 Average Budget Priority Financing Gap")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="2E75B6")
    row += 1

    agg = aggregate_with_gini(df, 'Gap Gov exp Health on budget')
    row = add_dataframe_to_sheet(ws, agg, row, "Africa Level")

    for group in ['income', 'Subregion']:
        agg = aggregate_with_gini(df, 'Gap Gov exp Health on budget', group)
        row = add_dataframe_to_sheet(ws, agg, row, f"By {group}")

    # 3.2.3 Gini coefficient
    ws.cell(row=row, column=1, value="3.2.3 Inequality in Budget Priority")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="2E75B6")
    row += 1

    agg = aggregate_with_gini(df, 'Gov exp Health on budget')
    row = add_dataframe_to_sheet(ws, agg, row, "Africa Level - Gini & Range")

    for group in ['income', 'Subregion']:
        agg = aggregate_with_gini(df, 'Gov exp Health on budget', group)
        row = add_dataframe_to_sheet(ws, agg, row, f"By {group}")

    adjust_column_widths(ws)


def create_section_33(wb, df):
    """3.3 Financial Protection of Households."""
    ws = wb.create_sheet(title="3.3 Financial Protection")
    row = 1

    ws.cell(row=row, column=1, value="3.3 Financial Protection of Households")
    ws.cell(row=row, column=1).font = Font(bold=True, size=14, color="1F4E79")
    row += 2

    # 3.3.1 Countries with OOP BELOW the 20% benchmark (good financial protection)
    # Spec: "Number of African countries with share of OOP... below the benchmark"
    ws.cell(row=row, column=1, value="3.3.1 Countries with OOP Below 20% Benchmark (Good Financial Protection)")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="2E75B6")
    row += 1

    # Count countries with good protection (OOP < 20%)
    agg = count_by_condition(df, 'below_oop_benchmark')
    row = add_dataframe_to_sheet(ws, agg, row, "Africa Level")

    for group in ['income', 'Subregion']:
        agg = count_by_condition(df, 'below_oop_benchmark', group)
        row = add_dataframe_to_sheet(ws, agg, row, f"By {group}")

    # 3.3.2 Average financial protection gap
    ws.cell(row=row, column=1, value="3.3.2 Average Financial Protection Gap (Excess OOP)")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="2E75B6")
    row += 1

    agg = aggregate_with_gini(df, 'Exc Out-of-pocket on health exp')
    row = add_dataframe_to_sheet(ws, agg, row, "Africa Level")

    for group in ['income', 'Subregion']:
        agg = aggregate_with_gini(df, 'Exc Out-of-pocket on health exp', group)
        row = add_dataframe_to_sheet(ws, agg, row, f"By {group}")

    # 3.3.3 Gini coefficient
    ws.cell(row=row, column=1, value="3.3.3 Inequality in Out-of-Pocket Expenditure")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="2E75B6")
    row += 1

    agg = aggregate_with_gini(df, 'Out-of-pocket on health exp')
    row = add_dataframe_to_sheet(ws, agg, row, "Africa Level - Gini & Range")

    for group in ['income', 'Subregion']:
        agg = aggregate_with_gini(df, 'Out-of-pocket on health exp', group)
        row = add_dataframe_to_sheet(ws, agg, row, f"By {group}")

    # 3.3.4 Financial hardship
    ws.cell(row=row, column=1, value="3.3.4 Incidence of Financial Hardship")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="2E75B6")
    row += 1

    agg = aggregate_with_gini(df, 'financial hardship')
    row = add_dataframe_to_sheet(ws, agg, row, "Africa Level")

    for group in ['income', 'Subregion']:
        agg = aggregate_with_gini(df, 'financial hardship', group)
        row = add_dataframe_to_sheet(ws, agg, row, f"By {group}")

    adjust_column_widths(ws)


def create_section_34(wb, df):
    """3.4 Health Financing Structure."""
    ws = wb.create_sheet(title="3.4 Financing Structure")
    row = 1

    ws.cell(row=row, column=1, value="3.4 Health Financing Structure (Sources of Health Financing)")
    ws.cell(row=row, column=1).font = Font(bold=True, size=14, color="1F4E79")
    row += 2

    # 3.4.1 Government dominant
    ws.cell(row=row, column=1, value="3.4.1 Countries with Government Share > 50% (Dominant)")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="2E75B6")
    row += 1

    agg = count_by_condition(df, 'gov_dominant')
    row = add_dataframe_to_sheet(ws, agg, row, "Africa Level")

    for group in ['income', 'Subregion']:
        agg = count_by_condition(df, 'gov_dominant', group)
        row = add_dataframe_to_sheet(ws, agg, row, f"By {group}")

    # 3.4.2 - 3.4.6 Average shares
    financing_indicators = [
        ('Govern on health exp', '3.4.2 Average Government Share'),
        ('Voluntary Prepayments on health exp', '3.4.3 Average Voluntary Prepaid Insurance Share'),
        ('Out-of-pocket on health exp', '3.4.4 Average Out-of-Pocket Share'),
        ('Other Private on health exp', '3.4.5 Average Other Private Share'),
        ('External on health exp', '3.4.6 Average Development Partners Share')
    ]

    for col, label in financing_indicators:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="2E75B6")
        row += 1

        agg = aggregate_with_gini(df, col)
        row = add_dataframe_to_sheet(ws, agg, row, "Africa Level")

        for group in ['income', 'Subregion']:
            agg = aggregate_with_gini(df, col, group)
            row = add_dataframe_to_sheet(ws, agg, row, f"By {group}")

    adjust_column_widths(ws)


def create_section_34a(wb, df):
    """3.4a Government Health Expenditure < 5% of GDP."""
    ws = wb.create_sheet(title="3.4a Gov Health Exp < 5% GDP")
    row = 1

    ws.cell(row=row, column=1, value="3.4a Number of African Countries with Government Share of Health Expenditure Less Than 5% of GDP")
    ws.cell(row=row, column=1).font = Font(bold=True, size=14, color="1F4E79")
    row += 2

    # Africa Level
    ws.cell(row=row, column=1, value="Countries with Government Health Expenditure < 5% of GDP")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="2E75B6")
    row += 1

    agg = count_by_condition(df, 'gov_health_below_5pct_gdp')
    row = add_dataframe_to_sheet(ws, agg, row, "Africa Level")

    # By Income Status
    agg = count_by_condition(df, 'gov_health_below_5pct_gdp', 'income')
    row = add_dataframe_to_sheet(ws, agg, row, "By Income Status")

    # By Sub-region
    agg = count_by_condition(df, 'gov_health_below_5pct_gdp', 'Subregion')
    row = add_dataframe_to_sheet(ws, agg, row, "By Sub-region")

    adjust_column_widths(ws)


def create_section_35(wb, df):
    """3.5 Health Outputs - UHC Index."""
    ws = wb.create_sheet(title="3.5 UHC Index")
    row = 1

    ws.cell(row=row, column=1, value="3.5 Health Outputs - Universal Health Coverage (UHC) Index")
    ws.cell(row=row, column=1).font = Font(bold=True, size=14, color="1F4E79")
    row += 2

    # 3.5.1 Average UHC
    ws.cell(row=row, column=1, value="3.5.1 Average UHC Index")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="2E75B6")
    row += 1

    agg = aggregate_with_gini(df, 'Universal health coverage')
    row = add_dataframe_to_sheet(ws, agg, row, "Africa Level")

    for group in ['income', 'Subregion']:
        agg = aggregate_with_gini(df, 'Universal health coverage', group)
        row = add_dataframe_to_sheet(ws, agg, row, f"By {group}")

    # 3.5.2 Countries with UHC < 50% OR below regional average
    # Spec: "UHC index of less than 50 per cent or lower than the average for the region"
    ws.cell(row=row, column=1, value="3.5.2 Countries with UHC Index Below 50% or Below Regional Average")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="2E75B6")
    row += 1

    agg = count_by_condition(df, 'uhc_below_50_or_avg')
    row = add_dataframe_to_sheet(ws, agg, row, "Africa Level")

    for group in ['income', 'Subregion']:
        agg = count_by_condition(df, 'uhc_below_50_or_avg', group)
        row = add_dataframe_to_sheet(ws, agg, row, f"By {group}")

    # 3.5.3 Gini coefficient
    ws.cell(row=row, column=1, value="3.5.3 Inequality in UHC Index")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="2E75B6")
    row += 1

    agg = aggregate_with_gini(df, 'Universal health coverage')
    row = add_dataframe_to_sheet(ws, agg, row, "Africa Level - Gini & Range")

    adjust_column_widths(ws)


def create_section_36(wb, df):
    """3.6 Health Outcomes - Mortality."""
    ws = wb.create_sheet(title="3.6 Health Outcomes")
    row = 1

    ws.cell(row=row, column=1, value="3.6 Health Outcomes - Neonatal mortality rate (NMR) and Maternal Mortality Ratio (MMR)")
    ws.cell(row=row, column=1).font = Font(bold=True, size=14, color="1F4E79")
    row += 1
    ws.cell(row=row, column=1, value="Note: Data uses Neonatal Mortality Rate (NMR) - deaths in the first 28 days of life per 1,000 live births.")
    ws.cell(row=row, column=1).font = Font(italic=True, size=10, color="FF0000")
    row += 2

    # 3.6.1 Average NMR 
    ws.cell(row=row, column=1, value="3.6.1 Average Neonatal mortality rate (per 1,000 live births)")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="2E75B6")
    row += 1

    agg = aggregate_with_gini(df, 'Neonatal mortality rate')
    row = add_dataframe_to_sheet(ws, agg, row, "Africa Level")

    for group in ['income', 'Subregion']:
        agg = aggregate_with_gini(df, 'Neonatal mortality rate', group)
        row = add_dataframe_to_sheet(ws, agg, row, f"By {group}")

    # 3.6.2 Average MMR
    ws.cell(row=row, column=1, value="3.6.2 Average Maternal Mortality Ratio (per 100,000 live births)")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="2E75B6")
    row += 1

    agg = aggregate_with_gini(df, 'Maternal mortality ratio')
    row = add_dataframe_to_sheet(ws, agg, row, "Africa Level")

    for group in ['income', 'Subregion']:
        agg = aggregate_with_gini(df, 'Maternal mortality ratio', group)
        row = add_dataframe_to_sheet(ws, agg, row, f"By {group}")

    # 3.6.3 Countries on course for NMR/NMR target
    # Spec: "reduce NMR to at least 12 per 1,000 live births" (meaning ≤ 12)
    ws.cell(row=row, column=1, value="3.6.3 Countries On Course to Reduce NMR to At Least 12 per 1,000 (NMR ≤ 12)")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="2E75B6")
    row += 1

    agg = count_by_condition(df, 'nmr_on_course')
    row = add_dataframe_to_sheet(ws, agg, row, "Africa Level")

    for group in ['income', 'Subregion']:
        agg = count_by_condition(df, 'nmr_on_course', group)
        row = add_dataframe_to_sheet(ws, agg, row, f"By {group}")

    # 3.6.4 Countries on course for MMR target
    # Spec: "reduce the MMR ratio to less than 70 per 100,000" (meaning < 70)
    ws.cell(row=row, column=1, value="3.6.4 Countries On Course to Reduce MMR to Less Than 70 per 100,000 (MMR < 70)")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="2E75B6")
    row += 1

    agg = count_by_condition(df, 'mmr_on_course')
    row = add_dataframe_to_sheet(ws, agg, row, "Africa Level")

    for group in ['income', 'Subregion']:
        agg = count_by_condition(df, 'mmr_on_course', group)
        row = add_dataframe_to_sheet(ws, agg, row, f"By {group}")

    adjust_column_widths(ws)


def create_section_37(wb, df):
    """3.7 Health Financing Dimensions and UHC Index."""
    ws = wb.create_sheet(title="3.7 Financing x UHC")
    row = 1

    ws.cell(row=row, column=1, value="3.7 Health Financing Dimensions and UHC Index")
    ws.cell(row=row, column=1).font = Font(bold=True, size=14, color="1F4E79")
    row += 2

    # 3.7.1 Threshold + UHC
    ws.cell(row=row, column=1, value="3.7.1 Countries Meeting Per Capita Threshold AND UHC Above 50th/75th Percentile")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="2E75B6")
    row += 1

    # Meeting threshold = NOT below threshold
    df['meets_threshold'] = 1 - df['below_threshold']

    for uhc_level, uhc_col in [('50th Percentile', 'uhc_above_p50'), ('75th Percentile', 'uhc_above_p75')]:
        ws.cell(row=row, column=1, value=f"UHC Above {uhc_level}")
        ws.cell(row=row, column=1).font = Font(italic=True, size=10)
        row += 1

        agg = cross_tabulation(df, 'meets_threshold', uhc_col)
        row = add_dataframe_to_sheet(ws, agg, row, "Africa Level")

        for group in ['income', 'Subregion']:
            agg = cross_tabulation(df, 'meets_threshold', uhc_col, group)
            row = add_dataframe_to_sheet(ws, agg, row, f"By {group}")

    # 3.7.2 Abuja + UHC
    ws.cell(row=row, column=1, value="3.7.2 Countries Meeting Abuja Target AND UHC Above 50th/75th Percentile")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="2E75B6")
    row += 1

    df['meets_abuja'] = 1 - df['below_abuja']

    for uhc_level, uhc_col in [('50th Percentile', 'uhc_above_p50'), ('75th Percentile', 'uhc_above_p75')]:
        ws.cell(row=row, column=1, value=f"UHC Above {uhc_level}")
        ws.cell(row=row, column=1).font = Font(italic=True, size=10)
        row += 1

        agg = cross_tabulation(df, 'meets_abuja', uhc_col)
        row = add_dataframe_to_sheet(ws, agg, row, "Africa Level")

    # 3.7.3 OOP + UHC
    # Countries with OOP below 20% benchmark (good financial protection) AND high UHC
    ws.cell(row=row, column=1, value="3.7.3 Countries with OOP Below 20% Benchmark AND UHC Above 50th/75th Percentile")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="2E75B6")
    row += 1

    for uhc_level, uhc_col in [('50th Percentile', 'uhc_above_p50'), ('75th Percentile', 'uhc_above_p75')]:
        ws.cell(row=row, column=1, value=f"UHC Above {uhc_level}")
        ws.cell(row=row, column=1).font = Font(italic=True, size=10)
        row += 1

        agg = cross_tabulation(df, 'below_oop_benchmark', uhc_col)
        row = add_dataframe_to_sheet(ws, agg, row, "Africa Level")

    adjust_column_widths(ws)


def create_section_38(wb, df):
    """3.8 Health Financing Dimensions and Health Outcomes."""
    ws = wb.create_sheet(title="3.8 Financing x Outcomes")
    row = 1

    ws.cell(row=row, column=1, value="3.8 Health Financing Dimensions and Health Outcomes (NMR and MMR)")
    ws.cell(row=row, column=1).font = Font(bold=True, size=14, color="1F4E79")
    row += 2

    # 3.8.1 Threshold + NMR
    ws.cell(row=row, column=1, value="3.8.1 Countries Meeting Per Capita Threshold AND On Course for NMR Target")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="2E75B6")
    row += 1

    df['meets_threshold'] = 1 - df['below_threshold']

    agg = cross_tabulation(df, 'meets_threshold', 'nmr_on_course')
    row = add_dataframe_to_sheet(ws, agg, row, "Africa Level")

    for group in ['income', 'Subregion']:
        agg = cross_tabulation(df, 'meets_threshold', 'nmr_on_course', group)
        row = add_dataframe_to_sheet(ws, agg, row, f"By {group}")

    # 3.8.2 Abuja + MMR
    ws.cell(row=row, column=1, value="3.8.2 Countries Meeting Abuja Target AND On Course for MMR Target")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="2E75B6")
    row += 1

    df['meets_abuja'] = 1 - df['below_abuja']

    agg = cross_tabulation(df, 'meets_abuja', 'mmr_on_course')
    row = add_dataframe_to_sheet(ws, agg, row, "Africa Level")

    for group in ['income', 'Subregion']:
        agg = cross_tabulation(df, 'meets_abuja', 'mmr_on_course', group)
        row = add_dataframe_to_sheet(ws, agg, row, f"By {group}")

    adjust_column_widths(ws)


def create_section_39(wb, df):
    """3.9 Health Financing Structure and UHC Index."""
    ws = wb.create_sheet(title="3.9 Structure x UHC")
    row = 1

    ws.cell(row=row, column=1, value="3.9 Health Financing Structure and UHC Index")
    ws.cell(row=row, column=1).font = Font(bold=True, size=14, color="1F4E79")
    row += 2

    # Create indicators for each financing source being dominant
    financing_sources = [
        ('Govern on health exp', 'gov_highest', '3.9.1 Government Share Dominant (> 50%)'),
        ('Voluntary Prepayments on health exp', 'vol_highest', '3.9.2 Voluntary Prepaid Highest'),
        ('Out-of-pocket on health exp', 'oop_highest', '3.9.3 Out-of-Pocket Highest'),
        ('Other Private on health exp', 'other_highest', '3.9.4 Other Private Highest'),
        ('External on health exp', 'ext_highest', '3.9.5 Development Partners Highest')
    ]

    # For government, use > 50% threshold
    df['gov_highest'] = df['gov_dominant']

    # For others, check if they are the highest component
    for col, flag_col, label in financing_sources[1:]:
        df[flag_col] = (df['dominant_financing_source'] == col).astype(float)
        df.loc[df['dominant_financing_source'].isna(), flag_col] = np.nan

    for col, flag_col, label in financing_sources:
        ws.cell(row=row, column=1, value=f"{label} AND UHC Above Percentiles")
        ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="2E75B6")
        row += 1

        for uhc_level, uhc_col in [('50th Percentile', 'uhc_above_p50'), ('75th Percentile', 'uhc_above_p75')]:
            ws.cell(row=row, column=1, value=f"UHC Above {uhc_level}")
            ws.cell(row=row, column=1).font = Font(italic=True, size=10)
            row += 1

            agg = cross_tabulation(df, flag_col, uhc_col)
            row = add_dataframe_to_sheet(ws, agg, row, "Africa Level")

    adjust_column_widths(ws)


def create_section_310(wb, df):
    """3.10 Health Financing Structure and Health Outcomes."""
    ws = wb.create_sheet(title="3.10 Structure x Outcomes")
    row = 1

    ws.cell(row=row, column=1, value="3.10 Health Financing Structure and Health Outcomes")
    ws.cell(row=row, column=1).font = Font(bold=True, size=14, color="1F4E79")
    row += 2

    financing_sources = [
        ('gov_highest', '3.10.1/6 Government Share Dominant (> 50%)'),
        ('vol_highest', '3.10.2/7 Voluntary Prepaid Highest'),
        ('oop_highest', '3.10.3/8 Out-of-Pocket Highest'),
        ('other_highest', '3.10.4/9 Other Private Highest'),
        ('ext_highest', '3.10.5/10 Development Partners Highest')
    ]

    outcomes = [
        ('nmr_on_course', 'NMR ≤ 12 per 1,000 live births'),
        ('mmr_on_course', 'MMR < 70 per 100,000 live births')
    ]

    for fin_col, fin_label in financing_sources:
        for out_col, out_label in outcomes:
            ws.cell(row=row, column=1, value=f"{fin_label} AND On Course for {out_label}")
            ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="2E75B6")
            row += 1

            agg = cross_tabulation(df, fin_col, out_col)
            row = add_dataframe_to_sheet(ws, agg, row, "Africa Level")

            for group in ['income', 'Subregion']:
                agg = cross_tabulation(df, fin_col, out_col, group)
                row = add_dataframe_to_sheet(ws, agg, row, f"By {group}")

    adjust_column_widths(ws)


def create_section_311(wb, df):
    """3.11 Fiscal Space and Macroeconomic Constraints."""
    ws = wb.create_sheet(title="3.11 Fiscal Space")
    row = 1

    ws.cell(row=row, column=1, value="3.11 Fiscal Space and Macroeconomic Constraints")
    ws.cell(row=row, column=1).font = Font(bold=True, size=14, color="1F4E79")
    row += 2

    # 3.11.1 Health spending elasticity - NOT AVAILABLE
    ws.cell(row=row, column=1, value="3.11.1 Health Spending Elasticity")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="2E75B6")
    row += 1
    ws.cell(row=row, column=1, value="[Data not available in current dataset - requires time-series regression analysis]")
    ws.cell(row=row, column=1).font = Font(italic=True, color="FF0000")
    row += 2

    # 3.11.2 Tax revenue
    ws.cell(row=row, column=1, value="3.11.2 Tax Revenue as % of GDP")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="2E75B6")
    row += 1

    agg = aggregate_with_gini(df, 'Tax Revenue per GDP')
    row = add_dataframe_to_sheet(ws, agg, row, "Africa Level")

    for group in ['income', 'Subregion']:
        agg = aggregate_with_gini(df, 'Tax Revenue per GDP', group)
        row = add_dataframe_to_sheet(ws, agg, row, f"By {group}")

    # 3.11.3 Health expenditure to GDP
    ws.cell(row=row, column=1, value="3.11.3 Health Expenditure as % of GDP")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="2E75B6")
    row += 1

    agg = aggregate_with_gini(df, 'Exp Health on GDP')
    row = add_dataframe_to_sheet(ws, agg, row, "Africa Level")

    for group in ['income', 'Subregion']:
        agg = aggregate_with_gini(df, 'Exp Health on GDP', group)
        row = add_dataframe_to_sheet(ws, agg, row, f"By {group}")

    # 3.11.5 Gross fixed capital formation
    ws.cell(row=row, column=1, value="3.11.5 Gross Fixed Capital Formation")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="2E75B6")
    row += 1

    gfcf_cols = [
        ('Gross fixed capital formation, as % of Gross domestic product (GDP)', 'GFCF as % of GDP'),
        ('Gross fixed capital formation, in current US$ per capita', 'GFCF Per Capita (Current USD)'),
        ('Gross fixed capital formation, in constant (2023) US$ per capita', 'GFCF Per Capita (Constant USD)')
    ]

    for col, label in gfcf_cols:
        if col in df.columns:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=1).font = Font(italic=True, size=10)
            row += 1

            agg = aggregate_with_gini(df, col)
            row = add_dataframe_to_sheet(ws, agg, row, "Africa Level")

    # Missing indicators note
    ws.cell(row=row, column=1, value="Missing Indicators:")
    ws.cell(row=row, column=1).font = Font(bold=True, color="FF0000")
    row += 1
    missing = [
        "3.11.4 Institutional health investment share - Not in dataset",
        "3.11.6 Foreign direct investment - Not in dataset",
        "3.11.7 Investment returns on health expenditure - Not in dataset"
    ]
    for item in missing:
        ws.cell(row=row, column=1, value=item)
        ws.cell(row=row, column=1).font = Font(italic=True, color="FF0000")
        row += 1

    adjust_column_widths(ws)


def create_summary_sheet(wb, df):
    """Create summary sheet."""
    ws = wb.create_sheet(title="Summary", index=0)
    row = 1

    ws.cell(row=row, column=1, value="Health Financing Gap Analysis for Africa")
    ws.cell(row=row, column=1).font = Font(bold=True, size=18, color="1F4E79")
    row += 1
    ws.cell(row=row, column=1, value="Statistical Product - Complete Data Tables")
    ws.cell(row=row, column=1).font = Font(italic=True, size=12)
    row += 2

    # Data coverage
    ws.cell(row=row, column=1, value="DATA COVERAGE")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="1F4E79")
    row += 1
    ws.cell(row=row, column=1, value=f"Total Countries: {df['location'].nunique()}")
    row += 1
    ws.cell(row=row, column=1, value=f"Year Range: {int(df['year'].min())} - {int(df['year'].max())}")
    row += 1
    ws.cell(row=row, column=1, value=f"Total Records: {len(df):,}")
    row += 2

    # Thresholds used
    ws.cell(row=row, column=1, value="THRESHOLDS APPLIED")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="1F4E79")
    row += 1
    thresholds = [
        "Per Capita Health Expenditure: LICs=$112, LMICs=$146, UMICs=$477",
        "Abuja Declaration: 15% of government budget",
        "Out-of-Pocket: Below 20% benchmark",
        "Government Dominant: >50% of health expenditure",
        "Neonatal Mortality Target: ≤12 per 1,000 live births",
        "Maternal Mortality Target: <70 per 100,000 live births",
        "UHC Index: 50% minimum threshold"
    ]
    for t in thresholds:
        ws.cell(row=row, column=1, value=f"   • {t}")
        row += 1
    row += 1

    # Sections
    ws.cell(row=row, column=1, value="REPORT SECTIONS")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="1F4E79")
    row += 1
    sections = [
        "3.1 Public Health Financing - Adequacy of public health financing gap",
        "3.2 Abuja Declaration - Budgetary priority assigned to health",
        "3.3 Financial Protection - Out-of-pocket expenditure protection",
        "3.4 Financing Structure - Sources of health financing",
        "3.4a Gov Health Exp < 5% GDP - Countries with Gov health exp < 5% of GDP",
        "3.5 UHC Index - Universal Health Coverage outcomes",
        "3.6 Health Outcomes - Neonatal and Maternal Mortality",
        "3.7 Financing x UHC - Cross-tabulation analysis",
        "3.8 Financing x Outcomes - Cross-tabulation analysis",
        "3.9 Structure x UHC - Financing sources and UHC",
        "3.10 Structure x Outcomes - Financing sources and mortality",
        "3.11 Fiscal Space - Macroeconomic constraints",
        "Threshold x UHC - Threshold categories by UHC >50%",
        "Threshold x NMR - Threshold categories by NMR (>12 vs ≤12)",
        "Threshold x MMR - Threshold categories by MMR (>70 vs ≤70)"
    ]
    for s in sections:
        ws.cell(row=row, column=1, value=f"   {s}")
        row += 1
    row += 1

    # Income groups
    ws.cell(row=row, column=1, value="COUNTRIES BY INCOME GROUP")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="1F4E79")
    row += 1
    income_counts = df.groupby('income')['location'].nunique()
    for income, count in sorted(income_counts.items()):
        ws.cell(row=row, column=1, value=f"   {income}: {count} countries")
        row += 1
    row += 1

    # Sub-regions
    ws.cell(row=row, column=1, value="COUNTRIES BY SUB-REGION")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="1F4E79")
    row += 1
    subregion_counts = df.groupby('Subregion')['location'].nunique()
    for subregion, count in sorted(subregion_counts.items()):
        ws.cell(row=row, column=1, value=f"   {subregion}: {count} countries")
        row += 1

    ws.column_dimensions['A'].width = 70


def adjust_column_widths(ws):
    """Adjust column widths for readability."""
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = min(max_length + 2, 40)


def create_threshold_crosstab_by_year(df, outcome_col, outcome_label):
    """Create cross-tabulation of threshold categories vs an outcome BY YEAR ONLY (no group disaggregation)."""
    # Filter for valid data
    valid_df = df[df['gov_exp_pc_thres'].notna() & df[outcome_col].notna()].copy()

    if len(valid_df) == 0:
        return None, None

    # Define all threshold categories in order
    threshold_cols = ['Below 50% of expenditure threshold', '50-74.9% of expenditure threshold',
                      '75-99.9% of expenditure threshold', 'Meet expenditure threshold target']

    # Number of countries by threshold category (by year only)
    count_df = valid_df.groupby(['year', outcome_col, 'gov_exp_pc_thres']).size().reset_index(name='count')
    count_pivot = count_df.pivot_table(
        index=['year', outcome_col],
        columns='gov_exp_pc_thres',
        values='count',
        fill_value=0
    ).reset_index()

    # Ensure all threshold columns exist (add missing ones with 0)
    for col in threshold_cols:
        if col not in count_pivot.columns:
            count_pivot[col] = 0

    # Add total countries column
    total_by_year = valid_df.groupby(['year', outcome_col]).size().reset_index(name='Number of countries')
    count_pivot = count_pivot.merge(total_by_year, on=['year', outcome_col], how='left')

    # Reorder columns
    count_pivot = count_pivot[['year', outcome_col, 'Number of countries'] + threshold_cols]

    # Sort by year
    count_pivot = count_pivot.sort_values(['year', outcome_col])

    # Average expenditure by threshold category (by year only)
    avg_df = valid_df.groupby(['year', outcome_col, 'gov_exp_pc_thres'])['Gov exp Health per capita'].mean().reset_index()
    avg_pivot = avg_df.pivot_table(
        index=['year', outcome_col],
        columns='gov_exp_pc_thres',
        values='Gov exp Health per capita'
    ).reset_index()

    # Ensure all threshold columns exist (add missing ones with 0)
    for col in threshold_cols:
        if col not in avg_pivot.columns:
            avg_pivot[col] = 0

    avg_pivot = avg_pivot.merge(total_by_year, on=['year', outcome_col], how='left')
    avg_pivot = avg_pivot[['year', outcome_col, 'Number of countries'] + threshold_cols]

    # Sort by year
    avg_pivot = avg_pivot.sort_values(['year', outcome_col])

    # Replace any remaining NaN values with 0
    count_pivot = count_pivot.fillna(0)
    avg_pivot = avg_pivot.fillna(0)

    # Convert count columns to integers
    for col in ['Number of countries'] + threshold_cols:
        count_pivot[col] = count_pivot[col].astype(int)

    return count_pivot, avg_pivot


def create_threshold_crosstab(df, outcome_col, outcome_label, group_by):
    """Create cross-tabulation of threshold categories vs an outcome, disaggregated by group."""
    # Filter for valid data
    valid_df = df[df['gov_exp_pc_thres'].notna() & df[outcome_col].notna()].copy()

    if len(valid_df) == 0:
        return None, None

    # Define all threshold categories in order
    threshold_cols = ['Below 50% of expenditure threshold', '50-74.9% of expenditure threshold',
                      '75-99.9% of expenditure threshold', 'Meet expenditure threshold target']

    # Define sort order for income and subregion
    income_order = {'Low': 0, 'Lower-middle': 1, 'Upper-middle': 2}
    subregion_order = {'Central Africa': 0, 'Eastern Africa': 1, 'Central Africa': 2,
                       'Northern Africa': 3, 'Southern Africa': 4, 'Western Africa': 5}

    # Number of countries by threshold category
    count_df = valid_df.groupby(['year', group_by, outcome_col, 'gov_exp_pc_thres']).size().reset_index(name='count')
    count_pivot = count_df.pivot_table(
        index=['year', group_by, outcome_col],
        columns='gov_exp_pc_thres',
        values='count',
        fill_value=0
    ).reset_index()

    # Ensure all threshold columns exist (add missing ones with 0)
    for col in threshold_cols:
        if col not in count_pivot.columns:
            count_pivot[col] = 0

    # Add total countries column
    total_by_group = valid_df.groupby(['year', group_by, outcome_col]).size().reset_index(name='Number of countries')
    count_pivot = count_pivot.merge(total_by_group, on=['year', group_by, outcome_col], how='left')

    # Reorder columns - all threshold columns included
    count_pivot = count_pivot[['year', group_by, outcome_col, 'Number of countries'] + threshold_cols]

    # Sort by group_by (income or subregion) then year
    if group_by == 'income':
        count_pivot['_sort'] = count_pivot[group_by].map(income_order)
    else:
        count_pivot['_sort'] = count_pivot[group_by].map(subregion_order)
    count_pivot = count_pivot.sort_values(['_sort', 'year', outcome_col]).drop('_sort', axis=1)

    # Average expenditure by threshold category
    avg_df = valid_df.groupby(['year', group_by, outcome_col, 'gov_exp_pc_thres'])['Gov exp Health per capita'].mean().reset_index()
    avg_pivot = avg_df.pivot_table(
        index=['year', group_by, outcome_col],
        columns='gov_exp_pc_thres',
        values='Gov exp Health per capita'
    ).reset_index()

    # Ensure all threshold columns exist (add missing ones with 0)
    for col in threshold_cols:
        if col not in avg_pivot.columns:
            avg_pivot[col] = 0

    avg_pivot = avg_pivot.merge(total_by_group, on=['year', group_by, outcome_col], how='left')
    # Reorder columns - all threshold columns included
    avg_pivot = avg_pivot[['year', group_by, outcome_col, 'Number of countries'] + threshold_cols]

    # Sort by group_by (income or subregion) then year
    if group_by == 'income':
        avg_pivot['_sort'] = avg_pivot[group_by].map(income_order)
    else:
        avg_pivot['_sort'] = avg_pivot[group_by].map(subregion_order)
    avg_pivot = avg_pivot.sort_values(['_sort', 'year', outcome_col]).drop('_sort', axis=1)

    # Replace any remaining NaN values with 0
    count_pivot = count_pivot.fillna(0)
    avg_pivot = avg_pivot.fillna(0)

    # Convert count columns to integers
    for col in ['Number of countries'] + threshold_cols:
        count_pivot[col] = count_pivot[col].astype(int)

    return count_pivot, avg_pivot


def create_section_threshold_uhc(wb, df):
    """Create sheet for Threshold × UHC cross-tabulation."""
    ws = wb.create_sheet(title="Threshold x UHC")
    row = 1

    ws.cell(row=row, column=1, value="Threshold Categories × UHC (>50%)")
    ws.cell(row=row, column=1).font = Font(bold=True, size=14, color="1F4E79")
    row += 1
    ws.cell(row=row, column=1, value="Cross-tabulation of government health expenditure threshold categories with UHC")
    ws.cell(row=row, column=1).font = Font(italic=True, size=10)
    row += 2

    # === BY YEAR (Africa-wide) - First two tables ===
    # By Year - Number of Countries
    ws.cell(row=row, column=1, value="By Year - Number of Countries (Africa-wide)")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="2E75B6")
    row += 1

    count_df_year, avg_df_year = create_threshold_crosstab_by_year(df, 'uhc_above_50', 'UHC >50%')
    if count_df_year is not None:
        row = add_dataframe_to_sheet(ws, count_df_year, row, is_count_table=True)

    # By Year - Average Expenditure
    ws.cell(row=row, column=1, value="By Year - Average Gov Health Expenditure Per Capita (Africa-wide)")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="2E75B6")
    row += 1

    if avg_df_year is not None:
        row = add_dataframe_to_sheet(ws, avg_df_year, row)

    # === BY INCOME GROUP ===
    # By Income Group - Number of Countries
    ws.cell(row=row, column=1, value="By Income Group - Number of Countries")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="2E75B6")
    row += 1

    count_df, avg_df = create_threshold_crosstab(df, 'uhc_above_50', 'UHC >50%', 'income')
    if count_df is not None:
        row = add_dataframe_to_sheet(ws, count_df, row, is_count_table=True)

    # By Income Group - Average Expenditure
    ws.cell(row=row, column=1, value="By Income Group - Average Gov Health Expenditure Per Capita")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="2E75B6")
    row += 1

    if avg_df is not None:
        row = add_dataframe_to_sheet(ws, avg_df, row)

    # === BY SUBREGION ===
    # By Subregion - Number of Countries
    ws.cell(row=row, column=1, value="By Subregion - Number of Countries")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="2E75B6")
    row += 1

    count_df, avg_df = create_threshold_crosstab(df, 'uhc_above_50', 'UHC >50%', 'Subregion')
    if count_df is not None:
        row = add_dataframe_to_sheet(ws, count_df, row, is_count_table=True)

    # By Subregion - Average Expenditure
    ws.cell(row=row, column=1, value="By Subregion - Average Gov Health Expenditure Per Capita")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="2E75B6")
    row += 1

    if avg_df is not None:
        row = add_dataframe_to_sheet(ws, avg_df, row)

    adjust_column_widths(ws)


def create_section_threshold_imr(wb, df):
    """Create sheet for Threshold × NMR cross-tabulation."""
    ws = wb.create_sheet(title="Threshold x NMR")
    row = 1

    ws.cell(row=row, column=1, value="Threshold Categories × NMR (>12 vs ≤12)")
    ws.cell(row=row, column=1).font = Font(bold=True, size=14, color="1F4E79")
    row += 1
    ws.cell(row=row, column=1, value="Cross-tabulation of government health expenditure threshold categories with Neonatal mortality rate")
    ws.cell(row=row, column=1).font = Font(italic=True, size=10)
    row += 2

    # === BY YEAR (Africa-wide) - First two tables ===
    # By Year - Number of Countries
    ws.cell(row=row, column=1, value="By Year - Number of Countries (Africa-wide)")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="2E75B6")
    row += 1

    count_df_year, avg_df_year = create_threshold_crosstab_by_year(df, 'nmr_category', 'NMR')
    if count_df_year is not None:
        row = add_dataframe_to_sheet(ws, count_df_year, row, is_count_table=True)

    # By Year - Average Expenditure
    ws.cell(row=row, column=1, value="By Year - Average Gov Health Expenditure Per Capita (Africa-wide)")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="2E75B6")
    row += 1

    if avg_df_year is not None:
        row = add_dataframe_to_sheet(ws, avg_df_year, row)

    # === BY INCOME GROUP ===
    # By Income Group - Number of Countries
    ws.cell(row=row, column=1, value="By Income Group - Number of Countries")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="2E75B6")
    row += 1

    count_df, avg_df = create_threshold_crosstab(df, 'nmr_category', 'NMR', 'income')
    if count_df is not None:
        row = add_dataframe_to_sheet(ws, count_df, row, is_count_table=True)

    # By Income Group - Average Expenditure
    ws.cell(row=row, column=1, value="By Income Group - Average Gov Health Expenditure Per Capita")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="2E75B6")
    row += 1

    if avg_df is not None:
        row = add_dataframe_to_sheet(ws, avg_df, row)

    # === BY SUBREGION ===
    # By Subregion - Number of Countries
    ws.cell(row=row, column=1, value="By Subregion - Number of Countries")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="2E75B6")
    row += 1

    count_df, avg_df = create_threshold_crosstab(df, 'nmr_category', 'NMR', 'Subregion')
    if count_df is not None:
        row = add_dataframe_to_sheet(ws, count_df, row, is_count_table=True)

    # By Subregion - Average Expenditure
    ws.cell(row=row, column=1, value="By Subregion - Average Gov Health Expenditure Per Capita")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="2E75B6")
    row += 1

    if avg_df is not None:
        row = add_dataframe_to_sheet(ws, avg_df, row)

    adjust_column_widths(ws)


def create_section_threshold_mmr(wb, df):
    """Create sheet for Threshold × MMR cross-tabulation."""
    ws = wb.create_sheet(title="Threshold x MMR")
    row = 1

    ws.cell(row=row, column=1, value="Threshold Categories × MMR (>70 vs ≤70)")
    ws.cell(row=row, column=1).font = Font(bold=True, size=14, color="1F4E79")
    row += 1
    ws.cell(row=row, column=1, value="Cross-tabulation of government health expenditure threshold categories with Maternal Mortality Ratio")
    ws.cell(row=row, column=1).font = Font(italic=True, size=10)
    row += 2

    # === BY YEAR (Africa-wide) - First two tables ===
    # By Year - Number of Countries
    ws.cell(row=row, column=1, value="By Year - Number of Countries (Africa-wide)")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="2E75B6")
    row += 1

    count_df_year, avg_df_year = create_threshold_crosstab_by_year(df, 'mmr_category', 'MMR')
    if count_df_year is not None:
        row = add_dataframe_to_sheet(ws, count_df_year, row, is_count_table=True)

    # By Year - Average Expenditure
    ws.cell(row=row, column=1, value="By Year - Average Gov Health Expenditure Per Capita (Africa-wide)")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="2E75B6")
    row += 1

    if avg_df_year is not None:
        row = add_dataframe_to_sheet(ws, avg_df_year, row)

    # === BY INCOME GROUP ===
    # By Income Group - Number of Countries
    ws.cell(row=row, column=1, value="By Income Group - Number of Countries")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="2E75B6")
    row += 1

    count_df, avg_df = create_threshold_crosstab(df, 'mmr_category', 'MMR', 'income')
    if count_df is not None:
        row = add_dataframe_to_sheet(ws, count_df, row, is_count_table=True)

    # By Income Group - Average Expenditure
    ws.cell(row=row, column=1, value="By Income Group - Average Gov Health Expenditure Per Capita")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="2E75B6")
    row += 1

    if avg_df is not None:
        row = add_dataframe_to_sheet(ws, avg_df, row)

    # === BY SUBREGION ===
    # By Subregion - Number of Countries
    ws.cell(row=row, column=1, value="By Subregion - Number of Countries")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="2E75B6")
    row += 1

    count_df, avg_df = create_threshold_crosstab(df, 'mmr_category', 'MMR', 'Subregion')
    if count_df is not None:
        row = add_dataframe_to_sheet(ws, count_df, row, is_count_table=True)

    # By Subregion - Average Expenditure
    ws.cell(row=row, column=1, value="By Subregion - Average Gov Health Expenditure Per Capita")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="2E75B6")
    row += 1

    if avg_df is not None:
        row = add_dataframe_to_sheet(ws, avg_df, row)

    adjust_column_widths(ws)


def main():
    """Main function to generate the complete Excel report."""
    print("=" * 70)
    print("GENERATING COMPLETE HEALTH FINANCING GAP STATISTICAL PRODUCT")
    print("=" * 70)

    print("\nLoading and processing data...")
    df = load_data()
    print(f"Loaded {len(df):,} records for {df['location'].nunique()} countries")
    print(f"Year range: {int(df['year'].min())} - {int(df['year'].max())}")

    wb = Workbook()
    wb.remove(wb.active)

    print("\nCreating sheets...")

    print("  - Summary")
    create_summary_sheet(wb, df)

    print("  - 3.1 Public Health Financing")
    create_section_31(wb, df)

    print("  - 3.2 Abuja Declaration")
    create_section_32(wb, df)

    print("  - 3.3 Financial Protection")
    create_section_33(wb, df)

    print("  - 3.4 Financing Structure")
    create_section_34(wb, df)

    print("  - 3.4a Gov Health Exp < 5% GDP")
    create_section_34a(wb, df)

    print("  - 3.5 UHC Index")
    create_section_35(wb, df)

    print("  - 3.6 Health Outcomes")
    create_section_36(wb, df)

    print("  - 3.7 Financing x UHC (Cross-tabulation)")
    create_section_37(wb, df)

    print("  - 3.8 Financing x Outcomes (Cross-tabulation)")
    create_section_38(wb, df)

    print("  - 3.9 Structure x UHC (Cross-tabulation)")
    create_section_39(wb, df)

    print("  - 3.10 Structure x Outcomes (Cross-tabulation)")
    create_section_310(wb, df)

    print("  - 3.11 Fiscal Space")
    create_section_311(wb, df)

    print("  - Threshold x UHC (New cross-tabulation)")
    create_section_threshold_uhc(wb, df)

    print("  - Threshold x NMR (New cross-tabulation)")
    create_section_threshold_imr(wb, df)

    print("  - Threshold x MMR (New cross-tabulation)")
    create_section_threshold_mmr(wb, df)

    output_file = os.path.join(OUTPUT_DIR, 'Health_Financing_Gap_Statistical_Product_FINAL.xlsx')
    wb.save(output_file)

    print("\n" + "=" * 70)
    print(f"REPORT SAVED: {output_file}")
    print("=" * 70)
    print(f"\nSheets created: {len(wb.sheetnames)}")
    for sheet in wb.sheetnames:
        print(f"   - {sheet}")

    return output_file


if __name__ == "__main__":
    main()
